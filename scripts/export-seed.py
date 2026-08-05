#!/usr/bin/env python3
"""Exporte data/KPI.xlsx → data/seed-from-excel.json"""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "KPI.xlsx"
OUT = ROOT / "data" / "seed-from-excel.json"

def rows(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, 7)]
        if vals[0] is None:
            continue
        out.append(vals)
    return out

def main():
    wb = load_workbook(XLSX, data_only=True)
    seed = {
        "year": 2026,
        "weeks": [],
        "automationsMetier": [],
        "automationsOdoo": [],
        "phishing": [],
        "maintenances": [],
        "ticketsByType": {},
        "ticketsByAssignee": {},
        "ticketsByRequester": {},
    }
    ws = wb["2026"]
    for r in range(5, 57):
        month, week = ws.cell(r, 3).value, ws.cell(r, 4).value
        if month is None or week is None:
            continue
        seed["weeks"].append({
            "year": 2026,
            "month": int(month),
            "week": int(week),
            "ticketsHorsSlaCloture": ws.cell(r, 5).value,
            "ticketsHorsSlaPriseEnCharge": ws.cell(r, 6).value,
            "demandesItHebdo": ws.cell(r, 11).value,
            "demandesNonResoluesHebdo": ws.cell(r, 13).value,
            "informations": ws.cell(r, 15).value or "",
            "reaction": ws.cell(r, 16).value or "",
        })
    for vals in rows(wb["Automatisations métiers"]):
        seed["automationsMetier"].append({
            "year": int(vals[0]), "month": int(vals[1]), "week": int(vals[2]),
            "explanation": vals[3] or "", "responsible": vals[4] or "",
        })
    for vals in rows(wb["Automatisations Odoo"]):
        seed["automationsOdoo"].append({
            "year": int(vals[0]), "month": int(vals[1]), "week": int(vals[2]),
            "explanation": vals[3] or "", "responsible": vals[4] or "",
        })
    for vals in rows(wb["Tests Phishing"]):
        seed["phishing"].append({
            "year": int(vals[0]), "month": int(vals[1]), "week": int(vals[2]),
            "explanation": vals[3] or "", "responsible": vals[4] or "",
            "failures": int(vals[5] or 0),
        })
    for vals in rows(wb["Maintenances Production"]):
        seed["maintenances"].append({
            "year": int(vals[0]), "month": int(vals[1]), "week": int(vals[2]),
            "explanation": vals[3] or "", "responsible": vals[4] or "",
        })
    ws = wb["Nombre tickets par type"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    week_cols = [(c, headers[c - 1]) for c in range(2, ws.max_column) if headers[c - 1] and str(headers[c - 1]).startswith("2026-S")]
    for r in range(2, ws.max_row):
        cat = ws.cell(r, 1).value
        if not cat or cat == "Total hebdomadaire":
            continue
        for c, label in week_cols:
            seed["ticketsByType"].setdefault(label, {})[cat] = int(ws.cell(r, c).value or 0)
    ws = wb["Nombre tickets par responsable"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    week_cols = [(c, headers[c - 1]) for c in range(2, ws.max_column) if headers[c - 1] and str(headers[c - 1]).startswith("2026-S")]
    for r in range(2, ws.max_row):
        person = ws.cell(r, 1).value
        if not person or person == "Total hebdomadaire":
            continue
        for c, label in week_cols:
            seed["ticketsByAssignee"].setdefault(label, {})[person] = int(ws.cell(r, c).value or 0)
    OUT.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} ({len(seed['weeks'])} weeks)")

if __name__ == "__main__":
    main()
